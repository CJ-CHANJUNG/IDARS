# -*- coding: utf-8 -*-
import openpyxl

wb = openpyxl.load_workbook(r'd:\CJ\Project Manager\IDARS\IDARS PJT\Data\projects\샘플2_20251231_072311\step1_import\step1_data.xlsx')
ws = wb.active

# Get headers
headers = [cell.value for cell in ws[1]]
print('Headers:', headers)

# Get data
rows = [[cell.value for cell in row] for row in ws.iter_rows(min_row=2)]
data = [dict(zip(headers, row)) for row in rows]

# Check specific slips
slips = ['94524511', '94533565', '94526996']
for slip in slips:
    row = next((r for r in data if str(r.get('Billing Document', '')) == slip), None)
    print(f'\n{slip}:')
    if row:
        print(f"  Amount: {row.get('Amount')}")
        print(f"  Quantity: {row.get('Quantity')}")
    else:
        print('  NOT FOUND')
